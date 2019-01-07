from openpyxl import Workbook
from openpyxl import load_workbook
import xml.dom.minidom
import os
book = Workbook()
sheet = book.create_sheet(0)
sheet["A1"] = 'Google Submission' 
sheet['B1'] ='Progress'
sheet['C1'] ='Pass'
sheet['D1'] ='Failed'
sheet['E1'] ='Models'
sheet['F1'] ='Comment'
def getAttr(path,ele,attr):
   dom = xml.dom.minidom.parse(path + '/test_result.xml')
   root = dom.documentElement
   Ele = root.getElementsByTagName(ele)
   Attr = Ele[0].getAttribute(attr)
   return Attr
rootPath = '//10.75.10.81/Share/Dynamo2/CTS/V0.500'
dirs = os.listdir(rootPath)
path = rootPath + '/' + dirs[0] + '/results'
subDir = os.listdir(path)
subPath = path + '/' + subDir[0]
passCounts = []
failedCounts = []
Models = []
for i in range(len(dirs)):
    testPath = rootPath + '/' + dirs[i-1] + '/results'
    print('now we are in ' + str(testPath))
    sheet['A'+str(i+1)] = dirs[i-1]
    sheet['A'+str(i+1)] = '100%'
    resultPath = testPath + '/' + os.listdir(testPath)[0]
    passCount = getAttr(resultPath,'Summary','pass')
    sheet['C'+str(i+1)] = passCount
    failedCount = getAttr(resultPath,'Summary','failed')
    sheet['D'+str(i+1)] = failedCount
    modelDone = getAttr(resultPath,'Summary','models_done')
    modelTotal = getAttr(resultPath,'Summary','models_total')
    sheet['E'+str(i+1)] = modelDone + '/' + modelTotal
    if failedCount == 0:
        comment = "pass"
    else:
        comment = failedCount + 'items sorry do not know how to list out in an easy way'
    sheet['F'+str(i+1)] = comment
book.save(r'D:\xtsReport.xlsx')